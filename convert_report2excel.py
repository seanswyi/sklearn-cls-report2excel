import os
from argparse import ArgumentParser, Namespace
from typing import Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm


def convert_report2excel(
    workbook: Workbook,
    report: Union[pd.DataFrame, dict[str, float]],
    sheet_name: str = "",
) -> Workbook:
    """Function to convert classification report to formatted Excel file.

    An openpyxl.Workbook object must first be created outside of the func and provided.
    The func will create a formatted sheet, add it to the provided Workbook, and return it.
    """
    # Formatting for header row and column.
    #   Header color is light gray here.
    header_color = PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    )
    header_font = Font(name="Arial", bold=True)
    default_font = Font(name="Arial", bold=False)

    # For conditional formatting.
    #   Higher = greener.
    start_color = Color(rgb="E67C73")  # White
    mid_color = Color(rgb="FFFFFF")  # Red
    end_color = Color(rgb="57BB8A")  # Green
    color_scale_rule = ColorScaleRule(
        start_type="num",
        start_value=0.0,
        start_color=start_color,
        mid_type="num",
        mid_value=0.6,
        mid_color=mid_color,
        end_type="num",
        end_value=1.0,
        end_color=end_color,
    )

    # Floating point precision.
    float_prec = "0.0000"

    if isinstance(report, dict):
        report = pd.DataFrame(report).T
        report.reset_index(inplace=True)

    if "index" in report.columns.values:
        report.rename(columns={"index": "class"}, inplace=True)
    else:
        report.rename(columns={"Unnamed: 0": "class"}, inplace=True)

    df_xl = dataframe_to_rows(df=report, index=False)

    worksheet = workbook.create_sheet(title=sheet_name)

    for row in df_xl:
        worksheet.append(row)

    # Outer boundary formatting.
    outer_border_top_row = 1
    outer_border_top_col = 1
    outer_border_bottom_row = report.shape[0] + 1
    outer_border_bottom_col = report.shape[1]

    top_left_corner = Border(top=Side(style="thin"), left=Side(style="thin"))
    top_right_corner = Border(top=Side(style="thin"), right=Side(style="thin"))
    bottom_left_corner = Border(left=Side(style="thin"), bottom=Side(style="thin"))
    bottom_right_corner = Border(right=Side(style="thin"), bottom=Side(style="thin"))
    top = Border(top=Side(style="thin"))
    bottom = Border(bottom=Side(style="thin"))
    left = Border(left=Side(style="thin"))
    right = Border(right=Side(style="thin"))

    top_left = f"{get_column_letter(outer_border_top_col)}" f"{outer_border_top_row}"
    bottom_right = (
        f"{get_column_letter(outer_border_bottom_col)}" f"{outer_border_bottom_row}"
    )
    for row in worksheet[top_left:bottom_right]:
        for cell in row:
            if cell.row == 1 and cell.column == 1:
                worksheet.cell(
                    row=cell.row, column=cell.column
                ).border = top_left_corner
            elif cell.row == 1 and cell.column == report.shape[0]:
                worksheet.cell(
                    row=cell.row, column=cell.column
                ).border = top_right_corner
            elif cell.row == report.shape[0] + 1 and cell.column == 1:
                worksheet.cell(
                    row=cell.row, column=cell.column
                ).border = bottom_left_corner
            elif cell.row == report.shape[0] + 1 and cell.column == report.shape[1]:
                worksheet.cell(
                    row=cell.row, column=cell.column
                ).border = bottom_right_corner
            elif cell.row in range(1, report.shape[0] + 1) and cell.column == 1:
                worksheet.cell(row=cell.row, column=cell.column).border = left
            elif (
                cell.row in range(1, report.shape[0] + 1)
                and cell.column == report.shape[1]
            ):
                worksheet.cell(row=cell.row, column=cell.column).border = right
            elif cell.row == 1 and cell.column in range(1, report.shape[1] + 1):
                worksheet.cell(row=cell.row, column=cell.column).border = top
            elif cell.row == report.shape[0] + 1 and cell.column in range(
                1, report.shape[1] + 1
            ):
                worksheet.cell(row=cell.row, column=cell.column).border = bottom
            else:
                continue

    # Header boundary.
    row = 1
    start_col = 1
    end_col = report.shape[1]

    left = f"{get_column_letter(start_col)}{start_col}"
    right = f"{get_column_letter(end_col)}{end_col}"

    for cell in worksheet[left:right][0]:
        existing_border = cell.border
        new_border = Border(
            left=existing_border.left,
            right=existing_border.right,
            top=existing_border.top,
            bottom=Side(style="thin"),
        )
        worksheet.cell(row=cell.row, column=cell.column).border = new_border

    # Support boundary.
    support_col = report.columns.get_loc("support") + 1

    start_row = 1
    end_row = report.shape[0] + 1

    top = f"{get_column_letter(support_col)}{start_row}"
    bottom = f"{get_column_letter(support_col)}{end_row}"

    for row in worksheet[top:bottom]:
        for cell in row:
            existing_border = cell.border
            new_border = Border(
                left=Side(style="thin"),
                right=existing_border.right,
                top=existing_border.top,
                bottom=existing_border.bottom,
            )
            worksheet.cell(row=cell.row, column=cell.column).border = new_border

    # Label boundary.
    end_row = report.shape[0] + 1

    for row in worksheet["A1":f"A{end_row}"]:
        for cell in row:
            existing_border = cell.border
            new_border = Border(
                left=existing_border.left,
                right=Side(style="thin"),
                top=existing_border.top,
                bottom=existing_border.bottom,
            )
            worksheet.cell(row=cell.row, column=cell.column).border = new_border

    # Thick avg divider.
    avg_row = report.shape[0] + 2

    try:
        if "accuracy" in report.iloc[:, 0].values:
            avg_row = report[report.iloc[:, 0] == "accuracy"].index[0] + 2
        else:
            avg_row = report[report.iloc[:, 0] == "micro avg"].index[0] + 2

        start = f"A{avg_row}"
        end = f"{get_column_letter(report.shape[1])}{avg_row}"

        for row in worksheet[start:end]:
            for cell in row:
                existing_border = cell.border
                new_border = Border(
                    left=existing_border.left,
                    right=existing_border.right,
                    top=Side(style="thick"),
                    bottom=existing_border.bottom,
                )
                worksheet.cell(row=cell.row, column=cell.column).border = new_border
    except IndexError:
        pass

    # Conditional formatting.
    gradient_grid_start_row = 1
    gradient_grid_start_col = report.columns.get_loc("precision") + 1
    gradient_grid_end_row = avg_row + 2
    gradient_grid_end_col = support_col - 1

    grid_range = (
        f"{get_column_letter(gradient_grid_start_col)}{gradient_grid_start_row}:"
        f"{get_column_letter(gradient_grid_end_col)}{gradient_grid_end_row}"
    )
    worksheet.conditional_formatting.add(
        range_string=grid_range, cfRule=color_scale_rule
    )

    # Format fonts.
    for row in worksheet[top_left:bottom_right]:
        for cell in row:
            if cell.row == 1 or cell.column == 1:
                worksheet.cell(row=cell.row, column=cell.column).font = header_font
                worksheet.cell(row=cell.row, column=cell.column).fill = header_color
            else:
                worksheet.cell(row=cell.row, column=cell.column).font = default_font

    # Format floating point precision.
    top_left = "B2"
    bottom_right = f"{get_column_letter(support_col - 1)}" f"{report.shape[0] + 1}"

    for row in worksheet[top_left:bottom_right]:
        for cell in row:
            worksheet.cell(row=cell.row, column=cell.column).number_format = float_prec

    # Fix formatting issue when `accuracy` is outputted.
    if "accuracy" in report["class"].values:
        precision_col = report.columns.get_loc("precision") + 1
        worksheet.cell(
            row=avg_row,
            column=precision_col,
        ).value = ""

        recall_col = report.columns.get_loc("recall") + 1
        worksheet.cell(row=avg_row, column=recall_col).value = ""

        support_value = worksheet.cell(row=avg_row + 1, column=support_col).value

        worksheet.cell(row=avg_row, column=support_col).value = support_value

        if "predicted" in report.columns:
            worksheet.cell(row=avg_row, column=support_col + 1).value = support_value

    return workbook


def main(args: Namespace) -> None:
    if args.report_filename:
        report_filepaths = [args.report_filename]
    else:
        report_files: str = os.listdir(args.report_dir)
        report_filepaths: list[str] = [
            os.path.join(args.report_dir, f) for f in report_files
        ]

    print("Report files:")
    for fp in report_filepaths:
        print(f"\t{fp}")

    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove default sheet.

    pbar = tqdm(
        iterable=report_filepaths,
        desc="Converting classification reports to formatted Excel files",
        total=len(report_filepaths),
    )
    for file in pbar:
        report = pd.read_csv(file)
        sheet_name = os.path.splitext(file.split("/")[-1])[0]
        workbook = convert_report2excel(workbook, report, sheet_name)

    print(f"New Workbook has a total of {len(workbook.worksheets)} Worksheets.")

    if args.report_filename:
        save_filename = f"{os.path.splitext(args.report_filename)[0]}_excel.xlsx"
    else:
        save_filename = "reports_formatted.xlsx"

    save_filepath = os.path.join(args.save_dir, save_filename)

    print(f"Saving in {save_filepath}")
    workbook.save(save_filepath)


if __name__ == "__main__":
    parser = ArgumentParser()

    parser.add_argument(
        "--report_dir",
        default="",
        type=str,
        help="Directory where all of the individual reports are.",
    )
    parser.add_argument(
        "--report_filename", default="", type=str, help="Path to single report file."
    )
    parser.add_argument(
        "--save_dir", default="", type=str, help="Directory to save Excel files."
    )

    args = parser.parse_args()

    if args.save_dir == "":
        if args.report_filename:
            args.save_dir = os.path.dirname(args.report_filename)
        else:
            args.save_dir = args.report_dir

    main(args)
